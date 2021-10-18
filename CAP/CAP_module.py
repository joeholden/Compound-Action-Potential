import pyabf
import matplotlib.pyplot as plt
import numpy as np
import os
import statistics
import itertools as it
from scipy import integrate
import operator
from pathlib import Path
from openpyxl import Workbook, load_workbook


def main():
    global error_present
    error_present = False

    def find_impulse(cap_file):
        """This function will return a 3-member tuple. The first
        element is a list containing the time indices of the delivered pulse.
        The second is the time-point in ms that the pulse was delivered. The
        last element is the stimulus duration"""

        cap_file.setSweep(sweepNumber=0, channel=1)
        max_pulse_value = max(cap_file.sweepY)
        time_indices_of_pulse = []

        for index, element in enumerate(cap_file.sweepY):
            if element > 0.1 * max_pulse_value:
                time_indices_of_pulse.append(index)

        pulse_duration_ms = len(time_indices_of_pulse) / (SAMPLING_FREQUENCY / 1000)
        time_of_spike_onset = (time_indices_of_pulse[0] / SAMPLING_FREQUENCY) * 1000
        return time_indices_of_pulse, time_of_spike_onset, pulse_duration_ms

    def find_baseline_noise(cap_file):
        global time_indices_of_pulse, time_of_spike_onset, pulse_duration_ms, average_noise, average_spread_from_mean, standard_dev
        time_indices_of_pulse, time_of_spike_onset, pulse_duration_ms = find_impulse(cap_file)
        cap_file.setSweep(sweepNumber=0, channel=0)
        average_noise = statistics.mean(cap_file.sweepY[0:(time_indices_of_pulse[0] - 100)])

        standard_dev = np.std(cap_file.sweepY[0:(time_indices_of_pulse[0] - 100)])

        baseline_peaks = []
        baseline_troughs = []

        for value in cap_file.sweepY[0:(time_indices_of_pulse[0] - 100)]:
            if value < average_noise:
                baseline_troughs.append(value)
            else:
                baseline_peaks.append(value)
        average_of_baseline_peaks = statistics.mean(baseline_peaks)
        average_of_baseline_troughs = statistics.mean(baseline_troughs)
        average_spread_from_mean = statistics.mean((average_of_baseline_peaks - average_noise,
                                                    average_noise - average_of_baseline_troughs))
        return average_noise, standard_dev

    def find_regions(cap_file):
        cap_file.setSweep(sweepNumber=0, channel=0)

        # classify non-noise as 3.5 sigma... ~100% of all true noise.
        # indices are still original

        non_noise = []
        for index, mv in enumerate(cap_file.sweepY):
            if mv > avg_noise + (3.5 * standard_dev) or mv < avg_noise - (3.5 * standard_dev):
                non_noise.append((index, mv))

        # Dynamically re-type non_noise to remove pre-pulse entries and most post-CAP entries. Reduces chance or error
        non_noise = [(index, t) for (index, t) in non_noise if time_indices_of_pulse[-1] < index < 30000]

        # Boolean list if point in non_noise is above or below noise level average. This lets us separate groups
        above_baseline = [True if j > avg_noise else False for i, j in non_noise]

        # Find regions of 5 sign-contiguous neighbors as a criteria for a real signal
        # Returned is a list of tuples. The first is an index value that is from the enumeratiion of list
        def find_ranges(lst, n):
            """Return ranges for `n` or more repeated values. (index, # repeated adjacent)"""
            groups = ((k, tuple(g)) for k, g in it.groupby(enumerate(lst), lambda x: x[-1]))
            repeated = (idx_g for k, idx_g in groups if len(idx_g) >= n)
            return ((sub[0][0], sub[-1][0]) for sub in repeated)

        regions = list(find_ranges(above_baseline, n=5))

        region_left_index = []
        for (m, n) in regions:
            region_left_index.append(non_noise[m][0])

        # //////////////////////////////////////////////////

        # Walk down a peak group to find the boundary points

        ops = {
            "<=": operator.le,
            ">=": operator.ge
        }

        def walk_down_to_baseline(op_function):
            """ Starts at the left point mid-peak/trough that triggered the 4 sigma condition.
            It then walks up/down in both directions until it reaches the baseline noise spot.
            Returns the boundary points. """

            global boundary_1, boundary_2
            at_boundary_1 = False
            at_boundary_2 = False

            current_coordinate = point
            while not at_boundary_1:
                test = cap_file.sweepY[current_coordinate]
                if op_function(test, avg_noise):
                    at_boundary_1 = True
                    boundary_1 = current_coordinate
                else:
                    current_coordinate -= 1

            current_coordinate = point
            while not at_boundary_2:
                test = cap_file.sweepY[current_coordinate]
                if op_function(test, avg_noise):
                    at_boundary_2 = True
                    boundary_2 = current_coordinate
                else:
                    current_coordinate += 1

            return boundary_1, boundary_2

        # Use the walk_down_to_baseline function
        boundaries_and_stats = []
        for point in region_left_index:
            if cap_file.sweepY[point] > avg_noise:
                peak_identity = 'peak'
                operator_function = ops["<="]
            else:
                peak_identity = 'trough'
                operator_function = ops[">="]
            (t1, t2) = walk_down_to_baseline(operator_function)

            a = calculate_peak_areas(cap_file, boundaries=(t1, t2, peak_identity), average_noise=avg_noise)
            latency = calculate_peak_latency(cap_file, boundaries=(t1, t2, peak_identity))
            boundaries_and_stats.append((t1, t2, peak_identity, a, latency))

        return boundaries_and_stats

    def calculate_peak_areas(cap_file, boundaries, average_noise):
        """Uses simpson's integration approximation to get area under the peak
        boundaries parameter is expected as a 3- member tuple (l_bound, r_bound, peak_identity)
        average_noise is expected as a float"""
        bounds = boundaries
        noise = average_noise
        x_array = cap_file.sweepX[bounds[0]:bounds[1]]
        y_array = cap_file.sweepY[bounds[0]:bounds[1]]
        base = np.full(x_array.shape, noise)
        if bounds[2] == 'peak':
            area = integrate.simpson(y_array - base, x_array)
        elif bounds[2] == 'trough':
            area = integrate.simpson(base - y_array, x_array)
        else:
            area = 0
            print('something went wrong')
        return area

    def calculate_peak_latency(cap_file, boundaries):
        bounds = boundaries
        # in seconds
        latency = (bounds[0] - time_indices_of_pulse[-1]) / (SAMPLING_FREQUENCY)
        return latency

    # /////////////////////////////////////////run below////////////////////////////////////////////

    avg_noise, avg_spread_from_mean = find_baseline_noise(cap_file)

    cap_file.setSweep(sweepNumber=0, channel=0)

    plt.figure()
    plt.plot(cap_file.sweepX, cap_file.sweepY, color='k', linewidth=0.6)
    avg_baseline_array = np.full(cap_file.sweepX.shape, avg_noise)
    plt.plot(cap_file.sweepX, avg_baseline_array, color='blue', linewidth=0.6)

    plt.xlim(.327, .332)
    plt.ylim(-100, 400)

    b = find_regions(cap_file)

    # Unpack peaks / troughs information
    try:
        p2l, p2r, p2i, p2a, p2lat = b[1]
        p1l, p1r, p1i, p1a, p1lat = b[0]
    except IndexError:
        print('No CAP signal detected')

    # Case where only 2 peaks detected
    try:
        p3l, p3r, p3i, p3a, p3lat = b[2]
        p4l, p4r, p4i, p4a, p3lat = b[3]
    except IndexError:
        pass

    try:
        # Impulse artifacts not detected
        if p1i == 'peak' and p2i == 'trough':
            plt.fill_between(cap_file.sweepX[p1l:p1r], cap_file.sweepY[p1l:p1r], avg_baseline_array[p1l:p1r],
                             where=cap_file.sweepY[p1l:p1r] - avg_baseline_array[p1l:p1r] > 0, color='#6b5b95',
                             alpha=.8)
            plt.fill_between(cap_file.sweepX[p2l:p2r], cap_file.sweepY[p2l:p2r], avg_baseline_array[p2l:p2r],
                             where=cap_file.sweepY[p2l:p2r] - avg_baseline_array[p2l:p2r] < 0, color='#feb236',
                             alpha=.8)
            plt.annotate(f'Peak area = {round(p1a, 4)}\nTrough area = {round(p2a, 4)}\nLatency = {round(p1lat, 5)}',
                         xy=(0.65, 0.85),
                         xycoords='axes fraction')

            sheet["C" + f"{row}"] = p1a
            sheet["D" + f"{row}"] = p2a
            sheet["E" + f"{row}"] = p1lat

        # Impulse trough is detected. Shifts over one for analysis
        elif p1i == 'trough' and p2i == 'peak':
            plt.fill_between(cap_file.sweepX[p2l:p2r], cap_file.sweepY[p2l:p2r], avg_baseline_array[p2l:p2r],
                             where=cap_file.sweepY[p2l:p2r] - avg_baseline_array[p2l:p2r] > 0, color='#6b5b95',
                             alpha=.8)
            plt.fill_between(cap_file.sweepX[p3l:p3r], cap_file.sweepY[p3l:p3r], avg_baseline_array[p3l:p3r],
                             where=cap_file.sweepY[p3l:p3r] - avg_baseline_array[p3l:p3r] < 0, color='#feb236',
                             alpha=.8)
            plt.annotate(f'Peak area = {round(p2a, 4)}\nTrough area = {round(p3a, 4)}\nLatency = {round(p2lat, 5)}',
                         xy=(0.65, 0.85),
                         xycoords='axes fraction')
            sheet["C" + f"{row}"] = p2a
            sheet["D" + f"{row}"] = p3a
            sheet["E" + f"{row}"] = p2lat

        # Can't detect undershoot
        elif p1i == 'peak' and p2i not in locals():
            plt.fill_between(cap_file.sweepX[p1l:p1r], cap_file.sweepY[p1l:p1r], avg_baseline_array[p1l:p1r],
                             where=cap_file.sweepY[p1l:p1r] - avg_baseline_array[p1l:p1r] > 0, color='#6b5b95',
                             alpha=.8)
            plt.annotate(f'Peak area = {round(p1a, 4)}\nTrough area = not found\nLatency = {round(p2lat, 5)}',
                         xy=(0.65, 0.85),
                         xycoords='axes fraction')
            error_present = True
            sheet["C" + f"{row}"] = p1a
            sheet["D" + f"{row}"] = ''
            sheet["E" + f"{row}"] = p1lat

        # No peaks detected or other case?
        else:
            error_present = True
            # print('error- check code')

    except NameError:
        error_present = True
        # print('no CAP detected above noise')

    plt.title(f'{cap_file_text}')
    plt.savefig(f'png_files/{cap_file_text}.png')
    plt.xlabel('time (s)')
    plt.ylabel('voltage (mV)')
    # plt.show()


# ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
abf_files = []
file_names_text = []
folder_path = 'C:/Users/joema/Desktop/CAP files/'
SAMPLING_FREQUENCY = 50000

for file in os.listdir(folder_path):
    abf = pyabf.ABF(folder_path + file)
    abf_files.append(abf)
    cap_file_t = Path(folder_path + file).stem
    file_names_text.append(cap_file_t)

try:
    workbook = load_workbook(filename="C:/Users/joema/Desktop/CAP_output.xlsx")
except FileNotFoundError:
    workbook = Workbook()

sheet = workbook.active
sheet["A1"] = 'File Name'
sheet["B1"] = 'Outcome Success'
sheet["C1"] = 'Peak Area'
sheet["D1"] = 'Trough Area'
sheet["E1"] = 'Latency'

global row
row = 2

for i in range(len(abf_files)):
    cap_file = abf_files[i]
    cap_file_text = file_names_text[i]
    sheet["A" + f"{row}"] = cap_file_text
    main()
    if (error_present in globals() or locals()) and error_present:
        sheet["B" + f"{row}"] = 'error'
    plt.close()
    row += 1
    workbook.save(filename="C:/Users/joema/Desktop/CAP_output.xlsx")


workbook.close()
